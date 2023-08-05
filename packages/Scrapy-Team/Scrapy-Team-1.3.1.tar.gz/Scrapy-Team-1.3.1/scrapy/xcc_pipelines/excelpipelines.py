from openpyxl import Workbook
try:
    from Scrapy_Prj.items import FactoryCatergoryItem, FactoryMaterialItem, FactoryBaikeItem, DetailDataplatformItem, DetailResouItem
except:
    from scrapy.xcc_items.factoryitems import FactoryCatergoryItem, FactoryMaterialItem, FactoryBaikeItem, DetailDataplatformItem, DetailResouItem
'''
author: tieyongjie
'''


class ExcelFactoryPipeline(object):
    def __init__(self):
        self.name = ''
        self.wb = Workbook()
        self.wb.create_sheet("ware_category")
        self.ws_0 = self.wb["ware_category"]
        self.ws_0.append(
            ['sources', 'url', 'level', 'category_name', 'category_id', 'p_category_name', 'p_category_id', 'create_time',
             'creator', 'brand_name', 'brand_id'])
        self.wb.create_sheet("ware_detail")
        self.ws_1 = self.wb["ware_detail"]
        self.ws_1.append(
            ['sources', 'url', 'title', 'category_name', 'category_id', 'list_json', 'packing', 'brand_id', 'brand_name', 'create_time',
             'creator', 'img_url', 'raw_img_url', 'pdf_url', 'raw_pdf_url', 'descs', 'spider_flag', 'layout_design', 'min_work_tp', 'max_work_tp', 'interior_structure'])
        self.wb.create_sheet("ware_wiki")
        self.ws_2 = self.wb["ware_wiki"]
        self.ws_2.append(
            ['sources', 'url', 'title', 'brand_name', 'brand_id', 'category_name', 'category_id', 'create_time',
             'creator', 'application', 'feature', 'standard', 'overview', 'description'])

    def process_item(self, item, spider):
        self.name = spider.name
        if isinstance(item, FactoryCatergoryItem):
            self.ws_0.append([item.get('sources'), item.get('url'), item.get('level'), item.get('category_name'),
                            item.get('category_id'), item.get('p_category_name'), item.get('p_category_id'),
                            item['create_time'], item['creator'], item['brand_name'], item['brand_id']])
        if isinstance(item, FactoryMaterialItem):
            self.ws_1.append([item.get('sources'), item.get('url'), item.get('title'), item.get('category_name'),
                            item.get('category_id'), item.get('list_json'), item.get('packing'), item['brand_id'], item['brand_name'],
                            item['create_time'], item['creator'], item.get('img_url'), item.get('raw_img_url'), item.get('pdf_url'), item.get('raw_pdf_url'), item.get('descs'), item.get('spider_flag'), item.get('layout_design'), item.get('min_work_tp'), item.get('max_work_tp'), item.get('interior_structure')])
        if isinstance(item, FactoryBaikeItem):
            self.ws_2.append([item.get('sources'), item.get('url'), item.get('title'), item['brand_name'],
                              item['brand_id'], item.get('category_name'), item.get('category_id'),
                              item['create_time'], item['creator'], item.get('application'), item.get('feature'),
                              item.get('standard'), item.get('overview'), item.get('description')])
        return item

    def __del__(self):
        # 调用__del__() 销毁对象，释放其空间
        self.wb.save('excels/{}.xlsx'.format(self.name))


class ExcelDataplatformPipeline(object):
    def __init__(self):
        self.name = ''
        self.wb = Workbook()
        self.wb.create_sheet("ware_detail_dataplatform")
        self.ws = self.wb["ware_detail_dataplatform"]
        self.ws.append(
            ['sources', 'url', 'title', 'category_name', 'category_id', 'list_json', 'container_json', 'packing', 'brand_id', 'brand_name', 'create_time',
             'creator', 'img_url', 'raw_img_url', 'pdf_url', 'raw_pdf_url', 'descs', 'min_work_tp', 'max_work_tp'])

    def process_item(self, item, spider):
        self.name = spider.name
        if isinstance(item, DetailDataplatformItem):
            self.ws.append([item.get('sources'), item.get('url'), item.get('title'), item.get('category_name'),
                            item.get('category_id'), item.get('list_json'), item.get('container_json'), item.get('packing'), item['brand_id'], item['brand_name'],
                            item['create_time'], item['creator'], item.get('img_url'), item.get('raw_img_url'), item.get('pdf_url'), item.get('raw_pdf_url'), item.get('descs'), item.get('min_work_tp'), item.get('max_work_tp')])
        return item

    def __del__(self):
        # 调用__del__() 销毁对象，释放其空间
        self.wb.save('excels/{}.xlsx'.format(self.name))


class ExcelResouPipeline(object):
    def __init__(self):
        self.name = ''
        self.wb = Workbook()
        self.wb.create_sheet("ware_resou")
        self.ws = self.wb["ware_resou"]
        self.ws.append(
            ['sources', 'url', 'title', 'category_name', 'category_id', 'list_json', 'packing', 'brand_id', 'brand_name', 'spider_time', 'create_time',
             'creator', 'pdf_url', 'raw_pdf_url', 'img_url', 'raw_img_url', 'descs'])

    def process_item(self, item, spider):
        self.name = spider.name
        if isinstance(item, DetailResouItem):
            self.ws.append([item.get('sources'), item.get('url'), item.get('title'), item.get('category_name'),
                            item.get('category_id'), item.get('list_json'), item.get('packing'), item.get('brand_id'), item.get('brand_name'), item.get('spider_time'),
                            item.get('create_time'), item.get('creator'), item.get('pdf_url'), item.get('raw_pdf_url'), item.get('img_url'), item.get('raw_img_url'), item.get('descs')])
        return item

    def __del__(self):
        # 调用__del__() 销毁对象，释放其空间
        self.wb.save('excels/{}.xlsx'.format(self.name))
